import datetime as dt
from datetime import datetime
import pandas as pd
from objects.cashflow import *
from utils.data_loading import *
from objects.portfolio import init_portfolios, Portfolio
import logging
from datetime import datetime,date,timedelta
from scripts.generate_datev_mapping import generate_datev_ordernumber_map
import pathlib
import os
import more_itertools as it
import openpyxl
from config import config
import math
import utils.data_loading as dload
import re
import utils.utils as ut
import locale
from pandas.tseries.offsets import DateOffset
locale.setlocale(locale.LC_TIME, "de_DE")

BEGIN_DATE = "2019-01-01" 
END_DATE = datetime.now().strftime('%Y-%m-%d')
portlist = ['EEB1 (BVB)', 'EEB2 (ING)', 'EEB3 (DKB)', 'eB (SPK)', 'EGF1', 'EGS']

sfname = sf_engine("""
        SELECT o.order_number, o.sp_folder_name, o.rent_start_net
        FROM salesforce_cleaned.opportunities o
    """)
sfname['net_rent'] = round(sfname['rent_start_net'] * 1.19,2)

def collate_opos_xlsx(keep_spvs, file: str) -> pd.DataFrame:
    opos_dfs = pd.read_excel(file, sheet_name=None, header=2, dtype={"Konto": str,"Kunden-/Lief.-Nr.":str ,"Rechnungs-Nr.": str})
    keep_cols = ["Datum", "Konto", "Buchungstext","Kunden-/Lief.-Nr.", "Rechnungs-Nr.", "Belegfeld 2", "Betrag Soll", "Betrag Haben", "SPV Tag"]
    # keep_spvs = ['EEB1 (BVB)', 'EEB2 (ING)', 'EEB3 (DKB)', 'eB (SPK)', "EGF1"]
    spv_tag_dict = dict(zip(config.DATEV_SPV, config.DATEV_SPV_NEW))
    # Instead of iterating -- can use replace function ?? -- Kandy
    for key, item in opos_dfs.items():
        item["SPV Tag"] = spv_tag_dict[key]
    collated_df = pd.concat(list(opos_dfs.values()), ignore_index=True)
    collated_df = collated_df.loc[collated_df["SPV Tag"].isin(keep_spvs), keep_cols].copy()
    collated_df["Saldo"] = collated_df["Betrag Soll"].subtract(collated_df["Betrag Haben"], fill_value=0)
    # Reading floating point numbers lead to approximation errors - so rounding it 2
    collated_df["Saldo"] = collated_df["Saldo"].round(2)
    return collated_df

def load_opos_cf_csv(port_list) -> pd.DataFrame:
    # Load relevant columns from Datev file and do basic data preproc
    opos_path = os.path.join(config.DATA_DIR, "opos")
    opos_files = [f for f in pathlib.Path(opos_path).iterdir() if not f.is_dir()]
    opos_files = sorted(opos_files, key = os.path.getmtime)
    logging.info("Using OPOS file %s\n", str(opos_files[-1]))
    logging.warning("Using OPOS file %s\n", str(opos_files[-1]))
    opos_cur = opos_files[-1]
    opos = collate_opos_xlsx(port_list, opos_cur)
    return opos

def load_datev_mapping() -> pd.DataFrame:
    # Use mapping file to match Datev Belegfeld to SF Order_Number
    mapping_subdir = os.path.join(config.DATA_DIR, "datev_mapping")
    mapping_files = [f for f in pathlib.Path(mapping_subdir).iterdir() if not f.is_dir()]
    mapping_files = sorted(mapping_files, key=os.path.getmtime)
    return pd.read_csv(mapping_files[-1], sep=';')


def preproc_opos(opos_df, port_list: list, dryrun=True, exclude_old_eb=True) -> pd.DataFrame:
    """

    """
    datev_df = opos_df[["Rechnungs-Nr.","Kunden-/Lief.-Nr.", "Belegfeld 2", "Datum", "Saldo", "SPV Tag", "Konto", "Buchungstext"]]
    datev_df.columns = ["belegfeld1",'order_number', "belegfeld2", "pay_date" ,"pay_amount", "spv_tag", "konto", "desc"]

    logging.warning(
        "Following %s Belegfeld1 fields in bank transactions are NULL:\n",
        str(datev_df.belegfeld1.isna().sum())
    )

    logging.warning(
        "Following %s rows in bank transactions are redundant:\n%s\n",
        str(datev_df.duplicated(['belegfeld1','order_number', 'konto','pay_date', 'pay_amount', 'desc'], False).sum()),
        str(datev_df[datev_df.duplicated(['belegfeld1','order_number', 'konto','pay_date', 'pay_amount', 'desc'], False)].index.values)
    )

    # according to Stephan Cordes (01.03.22), transactions on konto 1210 should be ignored for EGF1
    mask_for_removal = (datev_df.spv_tag == "EGF1") & (datev_df.konto == "1210")
    datev_df = datev_df[~mask_for_removal]

    datev_df = datev_df[datev_df.spv_tag.isin(port_list)].copy()
    datev_df.loc[:, "pay_date"] = pd.to_datetime(datev_df["pay_date"], dayfirst=True)

    # Load Buchungstext overrides
    correction_file = os.path.join(config.MANUAL_COR_DIR, "datev_manual_corrections_order_number.xlsx")
    buchung_correction = pd.read_excel(correction_file, sheet_name="buchungstext")
    buchung_correction.dropna(subset=["amended_desc"], inplace=True)
    buchung_correction = buchung_correction.rename(columns= {'delta_amount':'pay_amount'})
    buchung_correction = buchung_correction[["belegfeld1", "pay_date", "pay_amount",
                                             "spv_tag", "desc", "amended_desc"]]

    # Reading floating point numbers lead to approximation errors - so rounding it 2
    buchung_correction['pay_amount'] = buchung_correction.pay_amount.round(2)

    merged_df = datev_df.merge(buchung_correction, how="left")
    keep_rows = merged_df.amended_desc != "REMOVE"
    merged_df = merged_df[keep_rows].copy()
    update_rows = ~merged_df.amended_desc.isna()
    merged_df.loc[update_rows, "desc"] = merged_df.amended_desc[update_rows]
    datev_df = merged_df.drop("amended_desc", axis=1)

    #Fill Order number
    datev_df['lower_belegfeld1']=datev_df['belegfeld1'].str.lower()
    datev_df['order_number'] = (datev_df.groupby('lower_belegfeld1', group_keys=False)['order_number']
                            .apply(lambda x: x.bfill().ffill()))


    # load belegfeld1 to SF order_number mapping
    matched_order_num = load_datev_mapping()
    matched_order_num.loc[:,"lower_belegfeld1"] = matched_order_num.belegfeld1.str.lower()


    op_num2sf2 = sf_engine("""
        select o.order_number, o.sp_folder_name
        from salesforce_cleaned.opportunities o
    """)
    op_num2sf2.columns = ["order_number", "sp_folder_name"]

    merge_df = pd.merge(matched_order_num, op_num2sf2, how = 'left', on ='order_number')
    merge_df['sp_folder_name_x'] = merge_df['sp_folder_name_x'].fillna(merge_df['sp_folder_name_y'])
    merge_df = merge_df.drop('sp_folder_name_y',axis=1)
    merge_df = merge_df.rename(columns = {'sp_folder_name_x':'sp_folder_name'})

    merge1_df = merge_df.merge(op_num2sf2, on ='sp_folder_name',how ='left')
    merge1_df.loc[merge1_df['spv_tag'] == 'eB (SPK)', 'order_number_x'] = merge1_df['order_number_y']
    merge1_df['order_number_x'] = merge1_df['order_number_x'].fillna(merge1_df['order_number_y'])
    merge1_df = merge1_df.drop('order_number_y',axis=1)
    merge1_df = merge1_df.rename(columns = {'order_number_x':'order_number'})
    merge_df = merge1_df
    logging.warning(
        "Following %s Belegfeld1 fields in Datev mapping are redundant:\n%s\n",
        str(merge_df.sp_folder_name.duplicated(keep = False).sum()),
        str(merge_df.sp_folder_name.duplicated(keep = False).index.values)
    )

    logging.warning(
    "Following %s Belegfeld1 fields in Datev mapping are Null\n",
    str(merge_df.sp_folder_name.isna().sum()),
    )

    order_num_mapping = merge_df[["lower_belegfeld1","sp_folder_name", "order_number"]]\
        .drop_duplicates(subset=["lower_belegfeld1"])\
        .set_index("lower_belegfeld1")
    df_with_order_num = datev_df.join(order_num_mapping, on="lower_belegfeld1", rsuffix='_right')
    df_with_order_num['order_number'] = df_with_order_num['order_number'].fillna(df_with_order_num['order_number_right'])
    datev_matched = df_with_order_num.dropna(subset=["sp_folder_name"]).copy()
    logging.warning("Unmapped %s records with the following %s belegfeld1 values will be dropped:\n%s\n",
        str(df_with_order_num.sp_folder_name.isna().sum()),
        str(len(df_with_order_num.belegfeld1[df_with_order_num.sp_folder_name.isna()].unique())),
        str(df_with_order_num.belegfeld1[df_with_order_num.sp_folder_name.isna()].unique())
    )

    logging.warning(
        "Following %s Order Number fields in OPOS are Null\n",
        str(matched_order_num.sp_folder_name.isna().sum()),
    )
    # Save the bank transactions and analyse later
    fname = pd.Timestamp.today().strftime('transactions-not-considered.xlsx')
    fname = os.path.join(config.OUTPUT_DIR, fname)
    with pd.ExcelWriter(fname) as writer:
        datev_matched[datev_matched.order_number.isna()].to_excel(writer)

    datev_matched.pay_amount.replace(",", ".", inplace=True, regex=True)
    datev_matched.loc[:, "pay_amount"] = datev_matched.pay_amount.astype("float")

    # Define regex patterns for extracting payment due dates from Buchungstext

    date_pat_long = r'((?:' + '|'.join(ut.MonthsDE.all) + r')(?:.*?(?:20)?[12][0-9]))'

    yr_pat_long = r'((?:20)?[12][0-9])'
    mnth_pat_long = r'(?P<month>' + '|'.join(ut.MonthsDE.all) + r')'
    date_pat_short = r'([01][0-9]/(?:20)?[12][0-9])'

    # Extract due dates in text form, keep only the first match
    long_date_str = datev_matched.desc.str\
        .extractall(date_pat_long)\
        .xs(0, axis=0, level=1)
    long_date_str = long_date_str[0]
    # assert long_date_str.index.duplicated(False).sum() == 0, \
    #     "multiple date strings were parse, expected 1 per record"

    parsed_yr = long_date_str.str.findall(yr_pat_long)
    parsed_yr = parsed_yr.apply(lambda v: max(v) if v else np.NaN)
    parsed_yr.name = "year"
    parsed_yr = parsed_yr\
        .dropna()\
        .astype("int")
    parsed_mnth = long_date_str.str.extractall(mnth_pat_long)
    parsed_mnth.reset_index(1, inplace=True)
    parsed_mnth.loc[:, "month"] = parsed_mnth.month.apply(lambda x: ut.MonthsDE.text2num(x))
    parsed_date = parsed_mnth.join(parsed_yr)

    # split parsed_date by different representation in buchungstext (single date, list, and range of dates)
    multi_date_ix = parsed_date[parsed_date.match == 1].index
    single_date_mask = ~parsed_date.index.isin(multi_date_ix)
    single_date = parsed_date[single_date_mask].copy()

    date_range_ix = long_date_str[
        long_date_str.str.contains("-") & long_date_str.index.isin(multi_date_ix)
    ].index
    date_range = parsed_date.loc[date_range_ix].copy()
    date_list_ix = multi_date_ix[~multi_date_ix.isin(date_range.index)]
    date_list = parsed_date.loc[date_list_ix].copy()

    def process_date_range(df_group: pd.DataFrame):
        assert len(df_group) == 2, "Two rows are expected in the DataFrame"
        assert df_group["year"].iloc[0] == df_group["year"].iloc[1], \
            "All rows are expected to have the same year in the DataFrame"
        first_month = df_group["month"].iloc[0]
        last_month = df_group["month"].iloc[1]
        if last_month < first_month:
            first_month -= 12
        mnths_in_range = last_month - first_month + 1
        if mnths_in_range == 2:
            return df_group
        else:
            match = np.arange(mnths_in_range)
            month = np.arange(first_month, last_month + 1)
            yr = df_group["year"].iloc[0]
            year = np.array([yr if m > 0 else yr - 1 for m in month])
            month = [m if m > 0 else m + 12 for m in month]
            return pd.DataFrame(
                {"match": match, "month": month, "year": year}
            )

    def process_date_list(df_group: pd.DataFrame):
        if any(df_group.month.isin([11, 12])) & any(df_group.month.isin([1, 2])):
            h2_mask = df_group.month > 6
            df_group.loc[h2_mask, "year"] = df_group["year"].iloc[0] - 1
        if re.search(r'abzgl|abzueglich|\+ Rest |und Rest ', df_group.iloc[0, -1], flags=re.IGNORECASE):
            max_year = df_group[df_group.year == df_group.year.max()]
            max_month = max_year[max_year.month == max_year.month.max()]
            assert len(max_month) == 1
            return max_month
        else:
            return df_group

    # Handle Buchungstext that contains date ranges
    if date_range.empty:
        date_range_processed = pd.DataFrame()
    else:
        date_range_by_idx = date_range.groupby(date_range.index)
        date_range_processed = date_range_by_idx\
            .apply(process_date_range)\
            .reset_index(1, drop=True)

    # Handle lists of months
    if date_list.empty:
        date_list_processed = pd.DataFrame()
    else:
        date_list_by_idx = date_list\
            .join(datev_matched.desc)\
            .groupby(date_list.index)
        date_list_processed = date_list_by_idx\
            .apply(process_date_list)
        if date_list_processed.index.nlevels == 2:
            date_list_processed.reset_index(1, drop=True, inplace=True)

    processed_dates = pd.concat([
        single_date,
        date_range_processed,
        date_list_processed.iloc[:, :3]
    ])

    # How come we are we are very sure that NULL values will not be present -- Kandy
    assert processed_dates.isna().sum().sum() == 0,\
        "Partial match (only year or month) detected for long date match:\n" + \
        str(processed_dates.isna().sum())

    # Convert day month year columns to a single datetime
    processed_dates.rename({"match": "day"}, axis=1, inplace=True)
    processed_dates.loc[:, "day"] = 15  # delays are counted from the 15th
    processed_dates.loc[:, "year"] = processed_dates.year.apply(lambda y: y + 2000 if y < 100 else y)  # convert yy to yyyy
    long_date = pd.to_datetime(processed_dates)

    # Extract short form dates from Buchungstext
    # As of now, ignore the date range and list (at some point, need to include them) -- Kandy
    short_date = datev_matched.desc.str.extractall(date_pat_short)
    if short_date.shape[0] > 0:
        short_date = short_date.xs(0, axis=0, level=1)
        short_date = short_date[0]
        short_date = pd.to_datetime("15/"+short_date.str[:], dayfirst=True)

        # assert all(short_date.index.get_level_values(1) == 0), \
        #     "multiple short dates detected in one transaction"
        # short_date.reset_index(1, inplace=True, drop=True)
        # short_date = pd.to_datetime("05/"+short_date.loc[:, 0].str[:], dayfirst=True)

        # This also need to check and avoided -- rather than throwing error -- Kandy
        assert short_date.index.isin(long_date).sum() == 0,\
            "Detected records that matched the long as well as the short date pattern. Unexpected."
        # join parsed date to main dataframe
        all_date = short_date.append(long_date)
        all_date.name = "due_date"
    else:
        all_date = long_date
        all_date.name = "due_date"

    datev_matched = datev_matched.join(all_date)
    logging.warning("%s rows due date cannot be parsed from buchungstext", str(datev_matched.due_date.isna().sum()))
    # reduce pay amount proportionally where one payment spans several months
    frequency = datev_matched.index.value_counts()
    datev_matched.loc[:, "pay_amount"] = datev_matched.pay_amount/frequency

    # Classify cures from Buchungstext
    cure_rows = (
        (datev_matched.belegfeld2.str.lower() == "enpal") |
        (datev_matched.konto == '59999') | (
            datev_matched.desc.str.contains(r"\(Enpal\)|Erstattung.*Miete", regex=True) &
            datev_matched.belegfeld2.isnull()
        )
    ) & (datev_matched.pay_amount < 0)

    datev_matched["trans_type"] = np.where(cure_rows, "cure", "")

    # set all due dates to BOM in the output  # TODO: consider using 5th/15th instead of BOM in all dataframes?
    datev_matched.loc[:, "due_date"] = datev_matched["due_date"].astype("datetime64[M]")
    date_parse_failed = datev_matched[datev_matched.due_date.isna()].copy()
    datev_output_long = datev_matched\
        .loc[:, ["order_number","sp_folder_name", "belegfeld1", "spv_tag", "due_date", "pay_date",
                "pay_amount", "trans_type",
                "desc", "konto"]]\
        .dropna(subset=["due_date"])\
        .sort_values(["spv_tag", "order_number", "due_date", "pay_date"])\
        .set_index(["order_number", "spv_tag"], drop=True)
    if len(date_parse_failed) > 0:
        logging.warning(
            "Failed to parse dates for %i records: \n" +
            str(date_parse_failed.desc) + "\n",
            len(date_parse_failed)
        )

    date_parse_failed["problem"] = "due date"
    cols = ['belegfeld1', 'order_number', 'pay_date', 'pay_amount', 'spv_tag',
            'konto', 'desc', 'problem']

    # date_parse_failed should not have the same issue
    assert date_parse_failed.index.duplicated().sum() == 0

    manual_review = date_parse_failed[cols]
    manual_review["record_added_on"] = pd.Timestamp.today()
    manual_review.loc[manual_review.index.duplicated(False), "problem"] = "due date"
    manual_review.drop_duplicates(inplace=True)
    # change column order
    manual_review = manual_review[[
        'belegfeld1', 'pay_date', 'pay_amount', 'spv_tag',
        'konto', 'desc', 'order_number', 'problem', 'record_added_on'
    ]]

    # ignore older eB errors by default

    # append to file
    if not dryrun:
        filename = os.path.join(config.MANUAL_COR_DIR, "datev_manual_corrections_order_number.xlsx")
        book = openpyxl.load_workbook(filename)
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
            manual_review.to_excel(
                writer,
                sheet_name="buchungstext",
                startrow=writer.sheets["buchungstext"].max_row,
                index=False,
                header=False
            )

    return datev_output_long

#Load OPOS
def load_opos_cf(port_list: list, dryrun=True) -> pd.DataFrame:
    opos_df = load_opos_cf_csv(port_list)
    return preproc_opos(opos_df, port_list, dryrun)

def load_submissions(port_list: list, **kwargs) -> pd.DataFrame:
    return load_submissions_sql(port_list, **kwargs)

def load_submissions_sql(port_list: list, query_name) -> pd.DataFrame:
    """
    Create data source from csv extract of submitted/submittable opportunities
    :return: pd.Dataframe with the following columns:
        "order_number": str
        "spv_tag": str
        "first pay date": datetime
        "amount": numeric
        "schufa": str
    """
    # Retrieves all entries which are installed by fulfillment partner. Some of them may not be submitted
    # Why not retrieve only submitted installations -- Kandy

    # Run sql query to get submission data
    submissions = sf_engine(sql_queries[query_name])
    # Discard eB entries as these are loaded from csv later - why ?? Can we not use this as backup? -- Kandy

    logging.warning(
        "%s rows Order Numbers of Installations in SF2.0 contain NaN values\n",
        str(submissions.Order_Number__c.isna().sum())
    )
    submissions = submissions[~submissions.Order_Number__c.isna()]

    dup_sub = submissions.duplicated(subset='Order_Number__c', keep = False) & submissions['DateOfSubmission__c'].isna()
    submissions = submissions[~dup_sub]

    logging.warning(
        "Following %s Order Numbers of Installations in SF2.0 are redundant:\n%s\n",
        str(submissions.Order_Number__c.duplicated(False).sum()),
        str(submissions[submissions.Order_Number__c.duplicated(False)].Order_Number__c.values)
    )
    submissions = submissions[~submissions.Order_Number__c.duplicated(keep='first')]
    eb_sf_submissions = submissions[submissions.spv_tag__c.isin(('ezee_Benefit', 'ezee_Benefit_Discovergy'))]
    # Instead of removing, override with eb submissions CSV details
    # submissions = submissions[
    #     ~submissions.spv_tag__c.isin(('ezee_Benefit', 'ezee_Benefit_Discovergy'))
    # ].copy()

    if "eB (SPK)" in port_list:
        # eB requires CSV input!
        f_path = os.path.join(config.DATA_DIR, "sf_submissions", "eb_submissions.csv")
        submissions_eb = pd.read_csv(f_path, encoding="utf8", sep=";")
        submissions_eb.loc[:, "DateOfFirstRentalPayment__c"] = \
            pd.to_datetime(submissions_eb['DateOfFirstRentalPayment__c'], dayfirst=True)
        submissions_eb.loc[:, "DateOfSubmission__c"] = \
            pd.to_datetime(submissions_eb['DateOfSubmission__c'], dayfirst=True)
        submissions_eb.loc[:, "Rent_Monthly_Net__c"] = pd.to_numeric(
            submissions_eb["Rent_Monthly_Net__c"].str.replace(",", "."),
            downcast="unsigned"
        )
        submissions_eb = submissions_eb.drop("DateOfFirstRentalPayment__c", axis=1)

        logging.warning(
            "%s rows Order Numbers of Installations in eB Submissions File contain NaN values\n",
            str(submissions_eb.Order_Number__c.isna().sum())
        )
        submissions_eb = submissions_eb[~submissions_eb.Order_Number__c.isna()]

        logging.warning(
            "Following %s Order Numbers of Installations in eB Submissions File are redundant:\n%s\n",
            str(submissions_eb.Order_Number__c.duplicated(False).sum()),
            str(submissions_eb[submissions_eb.Order_Number__c.duplicated(False)].Order_Number__c.values)
        )
        submissions_eb = submissions_eb[~submissions_eb.Order_Number__c.duplicated(keep='first')]

        logging.warning(
            "eb SPK Submission entries in SF2.0 and Files are %s and %s respectively\n No. of matches is : %s\n",
            str(len(eb_sf_submissions)), str(len(submissions_eb)),
            str(eb_sf_submissions.Order_Number__c.isin(submissions_eb.Order_Number__c).sum())
        )

        submissions = pd.concat((submissions_eb, submissions), axis=0)

        logging.warning(
            "%s rows Order Numbers of Combined sources of Installations contain NaN values\n",
            str(submissions.Order_Number__c.isna().sum())
        )
        submissions = submissions[~submissions.Order_Number__c.isna()]

        logging.warning(
            "Following %s Order Numbers of Combined sources of Installations are redundant:\n%s\n",
            str(submissions.Order_Number__c.duplicated(False).sum()),
            str(submissions[submissions.Order_Number__c.duplicated(False)].Order_Number__c.values)
        )
        #Override the eB details in SalesForce with the details loaded from File
        submissions = submissions[~submissions.Order_Number__c.duplicated(keep='first')]

    # Rename SPV tags
    spv_tag_dict = dict(zip(config.EEB_SF + config.SPK_SF, config.EEB_NEW + config.SPK_NEW))
    submissions.replace(spv_tag_dict, inplace=True)

    # Rename columns
    new_col_names = dict(zip(config.SF_COL_OLD, config.SF_COL_NEW))
    submissions.rename(new_col_names, axis=1, inplace=True)
    submissions = submissions[submissions.spv_tag.isin(port_list)].copy()
    submissions.loc[:, "order_number"] = submissions.order_number.replace({
            '20200402_Lukić_00': '20200402_Lukic_00',
            '20191216_Groẞ_37': '20191216_Groß_37',
            '20200529_Topalović_21': '20200529_Topalovic_21'
        })

    # set index and change dtypes
    submissions.set_index(["order_number", "spv_tag"], drop=True, inplace=True)
    # downcasting may not be necessary -- kandy
    submissions.loc[:, "amount"] = pd.to_numeric(submissions["amount"], downcast="unsigned")
   

    # first pay date no longer part of SF2.0 query
    # submissions.loc[:, "first_pay_date"] = \
    #     pd.to_datetime(submissions["first_pay_date"])

    # converted from KeyError exception clause to if to be more explict
    # How does it handle null or empty values -- kandy
    # Ans: Filling with False values
    if "system_is_sold" in submissions.columns:
        submissions.loc[:, "system_is_sold"] = submissions["system_is_sold"].str.lower()
        submissions.loc[:, "system_is_sold"] = submissions.system_is_sold.map({"true": True, "false": False})

    if "submission_date" in submissions.columns:
        f_path = os.path.join(config.DATA_DIR, "dwh1", "submission-dates-from-sf1.xlsx")
        dwh1 = pd.read_excel(f_path)
    # fill missing Submission Dates from DWH1.0
    # Disregarding sf1 due to uninstallation
    # There are lack of 5117 submission date records in sf2 which exist in sf1
    # Adding sf1 via xlsx file.

    #     # Are we sure that all entries in SF1 is migrated to SF 2.0? This condition will always pass through -- Kandy
    #     dwh1 = sf_engine(old_sql_queries['submitted_submitable'])
    #     dwh1.rename(new_col_names, axis=1, inplace=True)
    #     dwh1 = dwh1[["order_number", "submission_date"]]
        # # some order numbers were changed in SF2.0 due to non-ASCII characters
        # If Nidhi does not fix issue then need to append more -- Kandy
        dwh1.loc[:, "order_number"] = dwh1.order_number.replace({
            '20200402_Lukić_00': '20200402_Lukic_00',
            '20191216_Groẞ_37': '20191216_Groß_37',
            '20200529_Topalović_21': '20200529_Topalovic_21'
        })

        logging.warning(
            "%s rows Order Numbers in SF1 contain NaN values\n",
            str(dwh1.order_number.isna().sum())
        )
        dwh1 = dwh1[~dwh1.order_number.isna()]

        logging.warning(
            "Following %s Order Numbers in SF1 are redundant:\n%s\n",
            str(dwh1.order_number.duplicated(False).sum()),
            str(dwh1[dwh1.order_number.duplicated(False)].order_number.values)
        )
        dwh1 = dwh1[~dwh1.order_number.duplicated(keep='first')]

        dwh1.set_index("order_number", inplace=True)
        submissions = submissions.join(dwh1, how="left", rsuffix="_1")
        na_mask = submissions.submission_date.isna()
        submissions.loc[na_mask, "submission_date"] = submissions.loc[na_mask, "submission_date_1"]
        submissions.drop("submission_date_1", axis=1, inplace=True)

        logging.warning(
            "%s rows Submission Dates in SF2 contain NaN values\n",
            str(na_mask.sum())
        )

        logging.warning(
            "%s rows Submission Dates in SF1 and SF2 contain NaN values\n",
            str(submissions.submission_date.isna().sum())
        )

   

    # added assertion error to break when there are missing submission dates
    # eB submission dates are loaded from csv, which does have a few missing values
    if query_name == "submitted":

        # As of now, skip entries whose submission date is not filled yet -- Kandy
        not_submitted_mask = submissions.submission_date.isna()
        not_submitted = submissions[not_submitted_mask]
        submissions = submissions[~not_submitted_mask]

        if not_submitted.shape[0] > 0:
            fname = pd.Timestamp.today().strftime('submitted_but_no_date-%Y%m%d.xlsx')
            fname = os.path.join(config.OUTPUT_DIR, fname)
            with pd.ExcelWriter(fname) as writer:
                not_submitted.to_excel(writer)

        missing_sub_date_mask = submissions.submission_date.isna() \
                                & (submissions.index.get_level_values(1) != 'eB (SPK)')
        missing_sub_date = submissions[missing_sub_date_mask].index
        assert missing_sub_date_mask.sum() == 0, "The following are missing submission dates: \n %s" % missing_sub_date

    # first pay date no longer part of SF2.0 query
    # first_pay_na = submissions.first_pay_date.isna()
    # n_na = first_pay_na.sum()
    # submissions.loc[:, "first_pay_date"].fillna(pd.Timestamp.today(), inplace=True)
    # if n_na > 0:
    #     logging.warning(
    #         "%i records have no first payment date, filling with today \n" +
    #         str(submissions.index[first_pay_na].values) + "\n",
    #         n_na)

    # Should this be a show stopping error? Or Ignore the order number, warn the user and continue -- Kandy
    mask_dup_ordnum = submissions.index.get_level_values(0).duplicated(False)
    assert mask_dup_ordnum.sum() == 0, \
        "multiple spv tags / amount detected for the same order number\n\n %s \n" \
        % (submissions.loc[mask_dup_ordnum, "id"])

    return submissions

def generate_opos_ordernumber_map(portlist, save_suffix="", dryrun=True):
    """
    Fill Null Order Number
    """

    datev_df = load_opos_cf_csv(portlist)
    datev_df = datev_df[["Konto", "Rechnungs-Nr.", "Kunden-/Lief.-Nr.", "Belegfeld 2", "SPV Tag", "Datum", "Buchungstext", "Saldo"]]
    datev_df.columns = ["konto", "belegfeld1", "order_number", "belegfeld2", "spv_tag", "pay_date", "desc", "pay_amount"]
    datev_df.loc[:, "pay_date"] = pd.to_datetime(datev_df["pay_date"], dayfirst=True)
    datev_df["konto_len"] = datev_df.konto.str.len()
    datev_df.sort_values(["konto_len", "pay_date"], na_position="first", inplace=True)

    datev_df['lower_belegfeld1']=datev_df['belegfeld1'].str.lower()
    datev_df['order_number'] = (datev_df.groupby('lower_belegfeld1', group_keys=False)['order_number']
                            .apply(lambda x: x.bfill().ffill()))

    # keep latest date
    unique_rows = datev_df\
        .drop_duplicates(subset=["konto", "belegfeld1", "order_number","spv_tag"], keep="last")\
        .reset_index(drop=True)

    # load manually provided belegfeld1 mapping
    manual_map_file = os.path.join(config.MANUAL_COR_DIR, "datev_manual_corrections_order_number.xlsx")
    manual_mapping = pd.read_excel(os.path.join(manual_map_file), sheet_name="missing_orderno", index_col=0).reset_index()
    # manual_mapping = manual_mapping.drop_duplicates(subset = ['order_number'],keep='last')
    manual_mapping = manual_mapping[manual_mapping['mapped_order_number'] != 'removed']
    # update match patterns with manual mapping
    merged_rows = pd.merge(unique_rows,manual_mapping, on = 'belegfeld1',how = 'left')
    merged_rows = merged_rows[merged_rows['mapped_order_number']!='no match']
    merged_rows['order_number'] = merged_rows['order_number'].fillna(merged_rows['mapped_order_number'])
    merged_rows = merged_rows[['konto_x', 'belegfeld1', 'order_number', 'belegfeld2', 'spv_tag_x',
        'pay_date', 'desc', 'pay_amount', 'konto_len','lower_belegfeld1']]
    merged_rows = merged_rows.rename(columns = {'konto_x':'konto','spv_tag_x':'spv_tag'})
    emp_order =  merged_rows[merged_rows['order_number'].isna()]

    logging.warning(
        "%i belegfeld values in the Datev extract were not matched:\n %s\n",
        len(emp_order['order_number'].unique()),
        str(emp_order['order_number'].unique())
    )

    # save belegfeld to order_number mapping file
    merged_rows.to_csv(
    os.path.join(config.DATA_DIR,
                    "datev_mapping",
                    "datev_mapping" + save_suffix + pd.Timestamp.today().strftime("%Y%m%d") + ".csv"),
    sep=';',
    index=False
    )

    # created dataframes for manual intervention
    mask_no_beleg = merged_rows.order_number.isna()
    df_no_orderno = merged_rows[mask_no_beleg]
    df_no_orderno["record_added_on"] = pd.Timestamp.today()
    df_no_orderno = df_no_orderno[['belegfeld1','konto','spv_tag','record_added_on']]
    if dryrun:
        fname = "datev_not_matched" + save_suffix + ".xlsx"
        fname = os.path.join(config.OUTPUT_DIR, fname)
        writer = pd.ExcelWriter(fname)
        df_no_orderno.to_excel(writer, sheet_name="missing_orderno", encoding="utf8", index=False)
        writer.close()
    else:
        # save records that require manual review by appending to a shared file
        filename = os.path.join(config.MANUAL_COR_DIR, "datev_manual_corrections_order_number.xlsx")
        book = openpyxl.load_workbook(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        df_no_orderno.to_excel(writer, sheet_name="missing_orderno", startrow=writer.sheets[name].max_row, index=False, header=False)
        writer.close()


generate_opos_ordernumber_map(portlist, dryrun=False)

opos_df = load_opos_cf(portlist, dryrun=False)
opos_df[['pay_date','due_date']] = opos_df[['pay_date','due_date']].apply(lambda x: pd.to_datetime(x).dt.strftime('%y/%m'))
cure = opos_df['trans_type'] == 'cure'
opos_df = opos_df[~cure].reset_index()
opos_df.loc[opos_df['sp_folder_name']=='P170762','order_number'] = 'ZZ000175021'


input  = pd.merge(opos_df,sfname, on='order_number', how='left')
input  = input[['order_number', 'spv_tag', 'belegfeld1', 'due_date', 'pay_date', 'pay_amount', 'trans_type','sp_folder_name_x', 'net_rent']]
input  = input.rename(columns={'sp_folder_name_x':'sp_folder_name','net_rent':'monthly_rent','pay_amount':'default_amount'})
input = input[['order_number','sp_folder_name','spv_tag','due_date','default_amount','monthly_rent']]
default_efiltered = input.sort_values(by=['sp_folder_name','due_date'])
default_efiltered= default_efiltered[['order_number','sp_folder_name','due_date','spv_tag','monthly_rent','default_amount']]

default_efiltered = default_efiltered.groupby(['order_number','sp_folder_name','due_date','spv_tag']).agg({'monthly_rent': 'max','default_amount':'sum'}).reset_index()
default_efiltered['month year'] = pd.to_datetime(default_efiltered['due_date'],format = '%y/%m').dt.strftime('%B %Y')

neg_amt = default_efiltered
neg_amt = neg_amt.groupby('order_number')['default_amount'].sum().reset_index()
neg_order = neg_amt[neg_amt['default_amount']<=0]['order_number']
neg_def = default_efiltered['order_number'].isin(neg_order)
default_efiltered = default_efiltered[~neg_def]

dict_spv = {
    'eB (SPK)':'Ezee Benefit', 
    'EEB1 (BVB)':'Enpal ezee Benefit GmbH', 
    'EEB3 (DKB)':'Enpal ezee Benefit III GmbH',
    'EEB2 (ING)':'Enpal ezee Benefit II GmbH', 
    'EGF1':'Enpal Green Future GmbH',
    'EGS': 'Enpal Green Solution'
}

dict_iban = {
    'eB (SPK)' : 'DE03 6535 1260 0134 1226 72',
    'EEB1 (BVB)' : 'DE14 1009 0000 2775 9870 09', 
    'EEB2 (ING)' : 'DE53 5002 1000 0018 1022 02',
    'EEB3 (DKB)' :'DE37 1203 0000 1020 8240 72',
    'EGF1' : 'DE28 2003 0000 0030 3160 95'
}

dict_bank = {
    'eB (SPK)' : 'Sparkasse Zollernalb',
    'EEB1 (BVB)':'Berliner Volksbank', 
    'EEB2 (ING)' : 'ING Bank',
    'EEB3 (DKB)' :'Deutsche Kreditbank',
    'EGF1' : 'UniCredit Bank AG'
}

dict_bic = {
    'eB (SPK)' : 'SOLADES1BAL',
    'EEB1 (BVB)' : 'BEVODEBB', 
    'EEB2 (ING)' : 'INGBDEFFXXX',
    'EEB3 (DKB)' :'BYLADEM1001',
    'EGF1' : 'HYVEDEMM300'
}

sfname = sf_engine("""
        SELECT o.order_number, o.opportunity_name as opportunity_name
        FROM salesforce_cleaned.opportunities o
        JOIN  salesforce_cleaned.cases c
        ON c.opportunity_id = o.opportunity_id 
            where
                c.record_type_name = 'Fulfilment'
                and c.opportunity_id = o.opportunity_id
                and c.is_submitted_to_spv = 'true'
                and c.is_deleted = 'false'   
    """)
sfname.columns = ["order_number", "opportunity_name"]

output_email = sfname.merge(default_efiltered, on='order_number',how = 'right')
output_email.loc[output_email['order_number'].isin(sfname.order_number),'refinanced'] = 'True'
output_email.loc[output_email['refinanced'].isna(),'refinanced'] = 'False'
logging.warning(
        "%s rows in output_email data contain NaN values for either Operation Number or/and Order Number\n%s\n",
        str(output_email.isna().sum().sum()),
        str(output_email[output_email['opportunity_name'].isna()])
)
output_email = output_email.drop_duplicates(keep='first')
output_email = output_email[['order_number','sp_folder_name','opportunity_name','spv_tag','month year','monthly_rent','default_amount','refinanced']]
output_email = output_email.rename(columns= {'order_number':'Order Number','sp_folder_name':'SP Folder Name','spv_tag':'SPV Tag','opportunity_name':'Opportunity Name'})



### Import to input data for Dashboard
num_cnt = output_email
num_cnt = num_cnt[num_cnt['default_amount'] >25]
num_cnt['count'] = output_email['SP Folder Name'].map(output_email['SP Folder Name'].value_counts())
num_cnt = num_cnt[num_cnt['count']>3]
num_cnt['month year'] = pd.to_datetime(num_cnt['month year'],format = '%B %Y')
num_cnt['month_plusone'] = num_cnt['month year'] + DateOffset(months = 1)
sep_list_tab = num_cnt[['SP Folder Name','month_plusone']]
merge_tab = sep_list_tab.merge(num_cnt, how = 'inner', left_on = ['SP Folder Name','month_plusone'], 
            right_on = ['SP Folder Name','month year']).drop('count',axis=1)
merge_tab['count'] = merge_tab['SP Folder Name'].map(merge_tab['SP Folder Name'].value_counts())
cons_list = merge_tab[merge_tab['count']>3]
cons_list['month year'] = cons_list['month year'] + DateOffset(months = 4)
cons_list = cons_list['SP Folder Name']
output_email.loc[output_email['SP Folder Name'].isin(cons_list),'4 consecutive'] = output_email['SP Folder Name']

opos_tab = output_email
opos_tab = opos_tab[opos_tab['refinanced'] == 'True']
# opos_tab = opos_tab[opos_tab['default_amount'] >25]
opos_tab['month year'] = pd.to_datetime(opos_tab['month year'],format = '%B %Y').dt.strftime('%Y/%m/%d')
opos_tab['Gesamtbetrag'] = opos_tab.groupby('Order Number')['default_amount'].cumsum()
neg_trans = opos_tab['Gesamtbetrag'] <=25
opos_tab = opos_tab[~neg_trans]
opos_tab['Gesamtbetrag'] = opos_tab['Gesamtbetrag'].fillna(0)
opos_tab['Anzahl der RLS'] = round(opos_tab['Gesamtbetrag']/opos_tab['monthly_rent'],2)
opos_tab = opos_tab.groupby(['SPV Tag','SP Folder Name','Order Number','4 consecutive','refinanced'],dropna=False).agg({'Anzahl der RLS':'max','Gesamtbetrag':'last'}).reset_index()
# opos_tab = opos_tab[['SP Folder Name','SPV Tag','month year','4 consecutive','refinanced','Anzahl der RLS']]
df_list = [opos_tab]
sheet_names = ['opos']
fname = datetime.now().strftime("opos_tab.xlsx")
fname = os.path.join(config.OUTPUT_DIR,'defaultreport_source',fname)
logging.info("writing to file")
with pd.ExcelWriter(fname) as writer:
    for i, name in zip(range(len(df_list)), sheet_names):
        df_list[i].to_excel(writer, sheet_name= name)


#Number assets:
total_assets_sf= sf_engine('''
        with pm_cases as (
        select opportunity_id, status, record_type_name
        from salesforce_cleaned.cases
        where status = 'Closed - Successful' 
        AND record_type_name = 'PortfolioManagement'
        ),
        fc_cases as (
        select *
        from salesforce_cleaned.fulfilment_cases
        where is_submitted_to_spv = 'true'
        --AND sales_call_1_open_timestamp is not null
        ),

        opportunities as (
        select *
        from salesforce_cleaned.opportunities
        where stage not in ('Closed - Lost', 'Closed - Unsuccessful')
        )

        SELECT  
            Row_Number() OVER(PARTITION BY op.opportunity_id ORDER By op.opportunity_id) as duplicates
            ,CAST([tranche_number] as int) [Tranche]
            ,op.opportunity_id
            ,op.[order_number] ID
            ,[sp_folder_name] [Customer ID]
            ,FORMAT(date_of_acceptance_appl_feed_in,'dd.MM.yyyy') [Date of grid connection]
            ,FORMAT (monthly_rent_net, 'c', 'de-DE') [Monthly rent]
            ,FORMAT (amount, 'c', 'de-DE') as [Purchase Price]
            ,spv_tag
            -- ,fc.case_id fulfillment_case_id
            -- ,portfolio_managment_closed_timestamp

        FROM pm_cases cs
        INNER JOIN fc_cases fc on cs.opportunity_id = fc.opportunity_id
        INNER JOIN opportunities op on cs.opportunity_id = op.opportunity_id
        ''')

total_assets = pd.DataFrame(total_assets_sf['spv_tag'].value_counts()).reset_index()
total_assets = total_assets.rename(columns = {'index':'spv_tag','spv_tag':'Asset'})
dict =  {"Enpal_ezee_Benefit":"EEB1 (BVB)",
        "Enpal_ezee_benefit_2":"EEB2 (ING)",
        "Enpal_ezee_Benefit_3":"EEB3 (DKB)",
        "enpal_green_future_1":"EGF1",
        "ezee_Benefit":"eB (SPK)"}

total_assets['spv_tag'] = total_assets['spv_tag'].replace(dict)

df_list = [total_assets]
sheet_names = ['asset']
fname = datetime.now().strftime("asset_tab.xlsx")
fname = os.path.join(config.OUTPUT_DIR,'defaultreport_source',fname)
logging.info("writing to file")
with pd.ExcelWriter(fname) as writer:
    for i, name in zip(range(len(df_list)), sheet_names):
        df_list[i].to_excel(writer, sheet_name= name)


## Normal processed opos\

output_email['SPV Tag'] = output_email['SPV Tag'].replace(dict_spv)
output_email['SPV_IBAN'] = output_email['SPV Tag'].replace(dict_iban)
output_email['SPV_Bank'] = output_email['SPV Tag'].replace(dict_bank)
output_email['SPV_BIC'] = output_email['SPV Tag'].replace(dict_bic)
output_email['Gesamtbetrag'] = output_email.groupby('Order Number')['default_amount'].cumsum()
neg_trans = output_email['Gesamtbetrag'] <=25
output_email = output_email[~neg_trans]
output_email['Monatsmiete'] = round(output_email['monthly_rent'],2)

output_email_month = output_email.groupby(['Order Number'])['month year'].apply(' & '.join)

# output_email['Anzahl der RLS'] = output_email['Order Number'].map(output_email['Order Number'].value_counts())
output_email['Gesamtbetrag'] = output_email['Gesamtbetrag'].fillna(0)
output_email['Anzahl der RLS'] = round(output_email['Gesamtbetrag']/output_email['monthly_rent'],2)

output_email = output_email.groupby(['Opportunity Name', 'SPV Tag','Order Number', 'SP Folder Name','SPV_IBAN', 'SPV_Bank', 'SPV_BIC','refinanced'],dropna=False).agg({'Anzahl der RLS':'max','Gesamtbetrag':'last','Monatsmiete':'max'}).reset_index()
output_email = output_email.merge(output_email_month, on = 'Order Number', how = 'left')

output_email['Frist'] = date.today() + timedelta(10)
output_email_fin = output_email[['Opportunity Name', 'SPV Tag','SPV_IBAN', 'SPV_Bank',
       'SPV_BIC','Order Number','SP Folder Name','Gesamtbetrag','Anzahl der RLS','Monatsmiete','month year','Frist','refinanced']]

output_email_fin = output_email_fin.rename(columns = {'month year':'Month Year'})


#Reason for defaulting
eeb_path = 'C:\\Users\\LeXuanAnhNguyen\\OneDrive - Enpal GmbH\\Documents\\reason\\EEB.xlsx'
eeb_df = pd.read_excel(eeb_path)
eeb_df = eeb_df[['Buchungstext','Verwendungszweck']]
eeb_df = eeb_df.rename(columns={'Buchungstext':'type','Verwendungszweck':'desc'})
eeb_df = eeb_df[eeb_df['type']=='RUECKLASTSCHRIFT']
eeb_df['order_number'] = eeb_df.desc.str.findall(r"KD\s\w+").str.get(0).str.slice(3,)
eeb_df['reason'] = eeb_df.desc.str.findall(r"(?:MS|AC|MD)\w+").str.get(0)
eeb_df = eeb_df[['order_number','reason']]

egf_path = "C:\\Users\\LeXuanAnhNguyen\\OneDrive - Enpal GmbH\\Documents\\reason\\EGF.xlsx"
egf_df = pd.read_excel(egf_path,sheet_name = 'Umsätze',dtype = {'GVC':str})
egf_df = egf_df[egf_df['Kontobezeichnung']=='EGF I - HVB - Proceeds Account  30316095'].reset_index()
egf_df = egf_df[egf_df['GVC']=='109']
egf_df = egf_df[['Verwendungszweck']]
egf_df = egf_df.rename(columns={'Verwendungszweck':'desc'})
egf_df['order_number'] = egf_df.desc.str.findall(r"KD\s\w+").str.get(0).str.slice(3,)
egf_df['reason'] = egf_df.desc.str.findall(r"(WIDERSPRUCH DURCH ZPFL|Sonstige Gruende|Widerspr bis 8 Woch|IBAN fehlerhaft)").str.get(0)
egf_df = egf_df[['order_number','reason']]

eeb2_path = 'C:\\Users\\LeXuanAnhNguyen\\OneDrive - Enpal GmbH\\Documents\\reason\\EEBII.xlsx'
eeb2_df = pd.read_excel(eeb2_path)
eeb2_df = eeb2_df[eeb2_df['Betrag in EUR']<0]
eeb2_df = eeb2_df[['Verwendungszweck']]
eeb2_df = eeb2_df.rename(columns={'Verwendungszweck':'desc'})
eeb2_df['order_number'] = eeb2_df.desc.str.findall(r"\w+CRED").str.get(0).str[:-4]
eeb2_df['reason'] = eeb2_df.desc.str.findall(r"(?:MS|AC|MD)\w{2}").str.get(0)
eeb2_df = eeb2_df[['order_number','reason']]
empty_ord = eeb2_df['order_number'].isna()
eeb2_df = eeb2_df[~empty_ord]

eeb3_path = 'C:\\Users\\LeXuanAnhNguyen\\OneDrive - Enpal GmbH\\Documents\\reason\\EEBIII.xlsx'
eeb3_df = pd.read_excel(eeb3_path)
eeb3_df = eeb3_df[eeb3_df['Buchungstext']=='Rückbelastung']
eeb3_df = eeb3_df[eeb3_df['Betrag (EUR)']<0]
eeb3_df = eeb3_df[['Verwendungszweck']]
eeb3_df = eeb3_df.rename(columns={'Verwendungszweck':'desc'}).reset_index()
eeb3_df['order_number'] = eeb3_df.desc.str.findall(r"KD\s\w+").str.get(0).str.slice(3,)
eeb3_df['reason'] = 'sonstiger grund'
eeb3_df = eeb3_df[['order_number','reason']]

# eb_path = 'C:\\Users\\LeXuanAnhNguyen\\OneDrive - Enpal GmbH\\Documents\\reason\\EB.xlsx'
# eb_df = pd.read_excel(eb_path)
# eb_df = eb_df[eb_df['Buchungstext']=='LS RUECKBELASTUNG']
# eb_df = eb_df[eb_df['Betrag']<0]
# eb_df = eb_df[['Mandatsreferenz','Verwendungszweck']]
# eb_df = eb_df.rename(columns={'Mandatsreferenz':'order_number','Verwendungszweck':'desc'}).reset_index()
# # eeb3_df['order_number'] = eeb3_df.desc.str.findall(r"KD\s\w+").str.get(0).str.slice(3,)
# eb_df['reason'] = eb_df.desc.str.findall(r"(Sonstige Gruende|Lastschriftwiderspruch durch den Zahlungspflichtigen|Kontonummer fehlerhaft|Konto aufgeloest)").str.get(0)
# eb_df = eb_df[['order_number','reason']]

default_reason = pd.concat([eeb_df,eeb2_df,eeb3_df,egf_df])
default_reason = default_reason.drop_duplicates(subset=['order_number'], keep='last')

output_fin = pd.merge(output_email_fin, default_reason, left_on = 'Order Number', right_on= 'order_number', how = 'left')

contract_changed_ord = sf_engine("""
    With open_operator_change as (
    select *
    from salesforce_cleaned.cases
    where sub_type in ('Change of Operator','death','House Selling','system sale')
    and is_closed = 'false'
    )

    SELECT op.order_number, ooc.*
    FROM portfoliomanagement.submitted_assets sa
    INNER JOIN salesforce_cleaned.opportunities op ON sa.opportunity_id = op.opportunity_id
    INNER JOIN open_operator_change ooc ON op.account_id = ooc.account_id
    """)

contract_changed_ord = contract_changed_ord[['order_number','sub_type']].loc[:,~contract_changed_ord[['order_number','sub_type']].columns.duplicated()]

output_fin_con = pd.merge(output_fin, contract_changed_ord, on='order_number',how='left')


fname = pd.Timestamp.today().strftime('%Y%m%d-Defaulted-customers.xlsx')
fname = os.path.join(config.OUTPUT_DIR,'datev_collection', fname)
with pd.ExcelWriter(fname) as writer:
    output_fin_con.to_excel(writer)
